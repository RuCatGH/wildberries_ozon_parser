import tkinter as tk
import json
import time
from tkinter import ttk
import os

import pandas as pd
import jmespath
import requests
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService 
from webdriver_manager.chrome import ChromeDriverManager 
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

from google_sheets import get_ozon_articles, get_wb_articles, get_guide
from headers import headers



def parse_wildberries(article: str, driver) -> list:
    """
    Парсит данные о продукте с веб-сайта Вайлдберриз

    Args:
        article (str): Артикул продукта
        driver: Драйвер Selenium для взаимодействия с веб-страницей

    Returns:
        list: Список с данными о продукте
    """
    response = requests.get(
        f'https://card.wb.ru/cards/detail?appType=1&curr=rub&dest=-1257786&regions=80,38,4,64,83,33,68,70,69,30,86,75,40,1,66,110,22,31,48,71,114&spp=31&nm={article}',
    )
    link = f'https://www.wildberries.ru/catalog/{article}/detail.aspx'
    price = jmespath.search("data.products[0].salePriceU", response.json())//100
    imtId = jmespath.search("data.products[0].root", response.json())
    total_valuation = jmespath.search("data.products[0].reviewRating", response.json())
    brand = jmespath.search("data.products[0].brand", response.json())
    price_without_discount = jmespath.search("data.products[0].priceU", response.json())//100
    

    discount1 = jmespath.search('data.products[0].extended.basicPriceU', response.json())//100
    

    discount2 = jmespath.search('data.products[0].extended.clientPriceU', response.json())//100

    wait = WebDriverWait(driver, 10)

    driver.get(f'https://www.wildberries.ru/catalog/{article}/feedbacks?imtId={imtId}')
    time.sleep(2)
    element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'product-line__price-now')))
    
    if driver.find_elements(By.CLASS_NAME, 'product-line__price-now')[-1].get_attribute('textContent') == 'Нет в наличии':
        availability = driver.find_elements(By.CLASS_NAME, 'product-line__price-now')[-1].text
    else:
        availability = 'В наличии'

    rating_elements = driver.find_elements(By.CLASS_NAME, 'feedback__rating')

    feedbacks = [int(class_value.get_attribute('class')[-1]) for class_value in rating_elements][:10]

    return [article, brand,link,availability, price,price_without_discount,discount1,discount2, total_valuation] + feedbacks

def parse_ozon(article: str, driver) -> list:
    """
    Парсит данные о продукте с веб-сайта Озон

    Args:
        article (str): Артикул продукта
        driver: Драйвер Selenium для взаимодействия с веб-страницей

    Returns:
        list: Список с данными о продукте
    """
    driver.get(f'https://www.ozon.ru/search/?text={article}&from_global=true')

    time.sleep(3)

    try:
        link = driver.find_elements(By.CLASS_NAME, 'tile-hover-target')[-1].get_attribute('href')


        link_json = link.split('https://www.ozon.ru')[1]
        
        driver.get(f'https://www.ozon.ru/api/entrypoint-api.bx/page/json/v2?url={link_json}&layout_container=pdpReviews&layout_page_index=2')
        time.sleep(1)
        json_data_html = driver.find_element(By.TAG_NAME, 'pre').text
        json_data = json.loads(json_data_html)

        valuation = []

        for key in json_data['widgetStates']:
            if key.startswith('webListReviews'):
                value = json_data['widgetStates'][key]
                json_feedbacks = json.loads(value)
                valuation = jmespath.search('reviews[:10].content.score', json_feedbacks)
        driver.get(link)

        json_price_original = json.loads(driver.find_element(By.XPATH, "//*[starts-with(@id, 'state-webPrice')]").get_attribute('data-state'))

        price_original = jmespath.search('originalPrice', json_price_original).replace('₽', '')
        price_with_card =  jmespath.search('cardPrice', json_price_original).replace('₽', '')

        json_product = json.loads(driver.find_element(By.CSS_SELECTOR, '[type="application/ld+json"]').get_attribute('textContent'))

        price = jmespath.search('offers.price', json_product)
        brand = jmespath.search('brand', json_product)

        try:
            total_valuation = jmespath.search('aggregateRating.ratingValue', json_product)
        except:
            total_valuation = 0
        availability = 'В наличии'
    except Exception as ex:
        print(ex)
        brand = None
        link = None
        price = None
        availability = 'Нет в наличии'
        price_with_card = None
        price_original = None
        total_valuation = None
        valuation = ['']

    return [article, brand, link, availability, price, price_with_card,price_original, total_valuation] + valuation
    



def parse_data_ozon() -> None:
    """
    Парсит данные с Вайлдберриз и Озон, сохраняет результаты в файлы Excel и выводит сообщение о завершении парсинга
    """
    try:
        ozon_workbook = Workbook()

        ws = ozon_workbook.active

        ws.append(['Артикул','Брэнд','Ссылка','Статус','Цена','Цена с озоной картой', 'Цена без скидки', 'Оценка за товар', 'Отзыв 1', 'Отзыв 2' , 'Отзыв 3', 'Отзыв 4', 'Отзыв 5', 'Отзыв 6', 'Отзыв 7', 'Отзыв 8', 'Отзыв 9', 'Отзыв 10'])

        options = webdriver.ChromeOptions()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--start-maximized")

        driver = uc.Chrome(options=options,service=ChromeService(ChromeDriverManager().install()))

        # Парсинг данных с OZON
        ozon_articles = get_ozon_articles()

        driver.get('https://www.ozon.ru/')
        time.sleep(7)
        for article in ozon_articles:
            data = parse_ozon(article, driver)
            ws.append(data)


        driver.close()
        driver.quit()

        ozon_workbook.save('ozon.xlsx')

        print("Парсинг завершен")
    except Exception as ex:
        ozon_workbook.save('ozon.xlsx')
        print(ex)

def parse_data_wb():
    try:
        wb_workbook = Workbook()
        ws = wb_workbook.active

        ws.append(['Артикул','Брэнд','Ссылка', 'Статус','Цена', 'Цена без скидки', 'Первая скидка цена','Вторая скидка цена','Оценка за товар', 'Отзыв 1', 'Отзыв 2' , 'Отзыв 3', 'Отзыв 4', 'Отзыв 5', 'Отзыв 6', 'Отзыв 7', 'Отзыв 8', 'Отзыв 9', 'Отзыв 10'])

        options = webdriver.ChromeOptions()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--start-maximized")

        driver = uc.Chrome(options=options,service=ChromeService(ChromeDriverManager().install()))

        # Парсинг данных с ВБ
        wb_articles = get_wb_articles()
        for article in wb_articles:
            data = parse_wildberries(article, driver)
            ws.append(data)

        # Сохранение результатов парсинга ВБ в файл wb.xlsx
        # wb_df = pd.DataFrame(wb_data, columns=['Артикул','Брэнд','Ссылка', 'Статус','Цена', 'Цена без скидки', 'Первая скидка цена','Вторая скидка цена','Оценка за товар', 'Отзыв 1', 'Отзыв 2' , 'Отзыв 3', 'Отзыв 4', 'Отзыв 5', 'Отзыв 6', 'Отзыв 7', 'Отзыв 8', 'Отзыв 9', 'Отзыв 10'])
        # wb_df.to_excel('wb.xlsx', index=False)
        
        driver.close()
        driver.quit()

        wb_workbook.save('wb.xlsx')

        print("Парсинг завершен")
    except Exception as ex:
        wb_workbook.save('wb.xlsx')
        print(ex)

def cross_reference_data() -> None: 
    """
    Загружает данные из файлов Excel, объединяет их и сохраняет результаты в файл Excel
    """

    # Проверка существования файла wb.xlsx
    if not os.path.isfile('wb.xlsx'):
        # Создание пустого файла wb.xlsx
        wb_workbook = Workbook()
        ws_wb = wb_workbook.active
        ws_wb.append(['Артикул'])
        wb_workbook.save('wb.xlsx')
        

    # Проверка существования файла ozon.xlsx
    if not os.path.isfile('ozon.xlsx'):
        # Создание пустого файла ozon.xlsx
        ozon_workbook = Workbook()
        ws_ozon = ozon_workbook.active
        ws_ozon.append(['Артикул'])
        ozon_workbook.save('ozon.xlsx')

    guide = get_guide()
    wb = pd.read_excel('wb.xlsx')
    ozon = pd.read_excel('ozon.xlsx')

    wb['Артикул'] = wb['Артикул'].astype(str)
    ozon['Артикул'] = ozon['Артикул'].astype(str)

    merged_df = pd.concat([wb, ozon], ignore_index=True)
    merged_df = pd.merge(guide, merged_df, on='Артикул')
    merged_df.to_excel('compare.xlsx', index=False)


def design() -> None:
    # Создаем графический интерфейс с использованием библиотеки tkinter
    window = tk.Tk()
    window.title("Парсер ВБ и OZON")
    window.geometry("300x200")
    window.configure(bg="#313841")

    style = ttk.Style()
    style.configure("Custom.TCheckbutton",
                    background="#313841",  # Цвет фона
                    foreground="white",  # Цвет текста
                    highlightcolor="white",  # Цвет обводки при наведении
                    relief=tk.FLAT,  # Убираем рамку
                    borderwidth=0,  # Убираем ширины границы
                    focuscolor="#6CC2F4",  # Цвет обводки при фокусе
                    focusthickness=2,  # Толщина обводки при фокусе
                    indicatorcolor="#6CC2F4"  # Цвет галочки
                    )

    button_style = {"fg": "white", "bd": 0, "relief": "flat", "font": ("Arial", 12, "bold"), "highlightthickness": 0, "highlightbackground": "#313841", "highlightcolor": "#313841", "overrelief": "flat"}

    parse_button = tk.Button(window, text="Спарсить ozon", command=parse_data_ozon, bg="#6CC2F4", **button_style)
    parse_button.place(x=50, y=30, width=200, height=40)

    parse_button = tk.Button(window, text="Спарсить wb", command=parse_data_wb, bg="#6CC2F4", **button_style)
    parse_button.place(x=50, y=80, width=200, height=40)

    compare_button = tk.Button(window, text="Сравнить", command=cross_reference_data, bg="#6CC2F4", **button_style)
    compare_button.place(x=50, y=130, width=200, height=40)


    window.mainloop()



if __name__ == '__main__':
    design()
    